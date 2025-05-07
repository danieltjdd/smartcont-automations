import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def formatar_ncm(ncm):
    ncm_str = re.sub(r'\D', '', str(ncm))
    ncm_str = ncm_str.zfill(8)
    if len(ncm_str) == 8:
        return f"{ncm_str[:4]}.{ncm_str[4:6]}.{ncm_str[6:]}"
    elif len(ncm_str) == 10:
        return ncm_str
    else:
        return ncm

def processar_pis_cofins(usuario_file, dominio_file=None):
    # Carrega domínio
    if dominio_file:
        dominio_df = pd.read_excel(dominio_file, dtype=str)
        dominio_df['ncm_formatado'] = dominio_df[dominio_df.columns[2]].apply(formatar_ncm)
        dominio_dict = dict(zip(dominio_df['ncm_formatado'], dominio_df[dominio_df.columns[1]]))
    else:
        dominio_dict = {}
    # Carrega usuário
    df = pd.read_excel(usuario_file, dtype=str)
    df.columns = [col.strip() for col in df.columns]
    df['NCM'] = df['NCM'].astype(str).str.zfill(8)
    colunas_esperadas = ['Número Nota', 'Cliente', 'CNPJ/CPF', 'Descricao', 'NCM', 'Vlr. Cont.', 'Valor ICMS']
    for col in colunas_esperadas:
        if col not in df.columns:
            raise Exception(f"Coluna obrigatória não encontrada: {col}")
    df['ncm_formatado'] = df['NCM'].apply(formatar_ncm)
    df['tributacao_full'] = df['ncm_formatado'].map(dominio_dict).fillna('')
    trib_split = df['tributacao_full'].str.split(';', n=2, expand=True)
    df['Tributação 1'] = trib_split[0].str.strip() if 0 in trib_split else ''
    df['Tributação 2'] = trib_split[1].str.strip() if 1 in trib_split else ''
    df['Tributação 3'] = trib_split[2].str.strip() if 2 in trib_split else ''
    df['Vlr. Cont.'] = pd.to_numeric(df['Vlr. Cont.'], errors='coerce').fillna(0)
    df = df.sort_values(['Tributação 1', 'Tributação 2', 'Tributação 3'])
    linhas = []
    for keys, grupo in df.groupby(['Tributação 1', 'Tributação 2', 'Tributação 3'], sort=False):
        for _, row in grupo.iterrows():
            linhas.append(row.to_dict())
        total = grupo['Vlr. Cont.'].sum()
        linha_total = {col: '' for col in df.columns}
        linha_total['Tributação 1'] = keys[0]
        linha_total['Tributação 2'] = keys[1]
        linha_total['Tributação 3'] = keys[2]
        linha_total['Vlr. Cont.'] = f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        linha_total['Descricao'] = 'TOTAL DO GRUPO'
        linhas.append(linha_total)
    colunas_finais = [
        'Número Nota', 'Cliente', 'CNPJ/CPF', 'Descricao', 'NCM', 'Vlr. Cont.', 'Valor ICMS',
        'Tributação', 'Cst de Entrada', 'Cst de Saida'
    ]
    df_organizado = pd.DataFrame(linhas, columns=df.columns)
    df_organizado = df_organizado.rename(columns={
        'Tributação 1': 'Tributação',
        'Tributação 2': 'Cst de Entrada',
        'Tributação 3': 'Cst de Saida'
    })
    for linha in linhas:
        if linha.get('Descricao') == 'TOTAL DO GRUPO':
            for col in colunas_finais:
                if col not in ['Tributação', 'Cst de Entrada', 'Cst de Saida', 'Vlr. Cont.', 'Descricao']:
                    linha[col] = ''
    df_organizado = df_organizado[colunas_finais]
    resumo = df.groupby(['Tributação 1', 'Tributação 2', 'Tributação 3'], as_index=False)['Vlr. Cont.'].sum()
    resumo['Vlr. Cont.'] = resumo['Vlr. Cont.'].apply(lambda x: f"R$ {x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
    df_organizado['__ordem__'] = df_organizado['Tributação'].apply(lambda x: 0 if str(x).strip().lower().startswith('tributado') else 1)
    df_organizado = df_organizado.sort_values(['__ordem__', 'Tributação']).drop(columns='__ordem__')
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f'relatorio_comparacao_{now}.xlsx'
    with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
        df_organizado.to_excel(writer, sheet_name='Organizado', index=False)
        resumo.to_excel(writer, sheet_name='Resumo', index=False)
    wb = openpyxl.load_workbook(nome_arquivo)
    ws = wb['Organizado']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
    moeda_cols = []
    for idx, col in enumerate(df_organizado.columns, 1):
        if col in ['Vlr. Cont.', 'Valor ICMS']:
            moeda_cols.append(idx)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in moeda_cols:
            cell = row[idx-1]
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.replace('R$', '').replace('.', '').replace(',', '').strip().isdigit()):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal='right')
    total_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
    total_font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        desc_cell = row[df_organizado.columns.get_loc('Descricao')]
        if desc_cell.value == 'TOTAL DO GRUPO':
            for cell in row:
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = ws['A2']
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    ws.auto_filter.ref = ws.dimensions
    wb.save(nome_arquivo)
    return nome_arquivo 