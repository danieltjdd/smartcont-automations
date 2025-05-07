import pandas as pd
import time
import openai
import os
from typing import List, Dict, Tuple
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ncm_validator.log'),
        logging.StreamHandler()
    ]
)

class NCMValidator:
    def __init__(self, api_key: str, batch_size: int = 10, wait_time: int = 21):
        openai.api_key = api_key
        self.batch_size = batch_size
        self.wait_time = wait_time
        
    def load_excel_files(self, user_file: str, ncm_file: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Carrega e prepara os arquivos Excel com as colunas corretas."""
        try:
            # Carrega planilha do usuário
            df_user = pd.read_excel(user_file)
            df_user = df_user[['NCM', 'Descricao']].copy()
            df_user.columns = ['NCM', 'Descricao']
            # Garante 8 dígitos com zeros à esquerda
            df_user['NCM'] = df_user['NCM'].astype(str).str.zfill(8)

            # Carrega tabela oficial de NCM
            df_ncm = pd.read_excel(ncm_file)
            df_ncm = df_ncm[['Codigo', 'Descricao']].copy()
            df_ncm.columns = ['Codigo', 'Descricao']
            # Coluna auxiliar: remove pontos e zeros à direita
            df_ncm['Codigo_Familia'] = df_ncm['Codigo'].astype(str).str.replace('.', '', regex=False).str.rstrip('0')
            return df_user, df_ncm
        except Exception as e:
            logging.error(f"Erro ao carregar arquivos Excel: {str(e)}")
            raise

    def buscar_descricao_concatenada(self, codigo: str, df_ncm: pd.DataFrame) -> str:
        """Busca a descrição concatenada da família e do NCM completo, reconhecendo códigos com pontos e zeros à direita."""
        desc_fam = None
        for tam in [4, 5, 6, 7]:
            familia = codigo[:tam].rstrip('0')
            match = df_ncm[df_ncm['Codigo_Familia'] == familia]
            if not match.empty:
                desc_fam = match.iloc[0]['Descricao']
                break
        if not desc_fam:
            desc_fam = "-- Família não encontrada"
        # Descrição do NCM completo
        codigo_ncm = codigo.lstrip('0')
        desc_ncm = df_ncm.loc[df_ncm['Codigo_Familia'] == codigo_ncm.rstrip('0'), 'Descricao']
        if desc_ncm.empty:
            desc_ncm = df_ncm.loc[df_ncm['Codigo'].str.replace('.', '', regex=False).str.zfill(8) == codigo, 'Descricao']
        desc_ncm = desc_ncm.values[0] if not desc_ncm.empty else "-- Outras"
        return f"{desc_fam} - {desc_ncm}"

    def validate_ncm_with_gpt(self, description: str, ncm: str) -> Dict:
        """Valida um NCM usando a API GPT."""
        try:
            prompt = f"""Analise a descrição do produto e o código NCM fornecido:\nDescrição: {description}\nNCM informado: {ncm}\n\nVerifique se o NCM está correto para o produto descrito. Responda apenas com:\n1. Se o NCM está correto ou não (CORRETO/INCORRETO)\n2. Se incorreto, sugira o NCM correto\n3. Justificativa breve da sua decisão\n\nResponda em formato JSON com as chaves: \"status\", \"ncm_sugerido\", \"justificativa\"\n"""
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3
            )
            return eval(response.choices[0].message.content)
        except Exception as e:
            logging.error(f"Erro na chamada da API GPT: {str(e)}")
            raise

    def process_in_batches(self, df_user: pd.DataFrame) -> pd.DataFrame:
        """Processa os registros em lotes."""
        results = []
        total_batches = (len(df_user) + self.batch_size - 1) // self.batch_size
        for i in range(0, len(df_user), self.batch_size):
            batch = df_user.iloc[i:i + self.batch_size]
            logging.info(f"Processando lote {(i//self.batch_size)+1} de {total_batches}")
            for _, row in batch.iterrows():
                try:
                    validation = self.validate_ncm_with_gpt(row['Descricao'], row['NCM'])
                    results.append({
                        'NCM': row['NCM'],
                        'Descricao': row['Descricao'],
                        'Status': validation['status'],
                        'NCM Sugerido': validation['ncm_sugerido'],
                        'Justificativa': validation['justificativa']
                    })
                except Exception as e:
                    logging.error(f"Erro ao processar item: {str(e)}")
                    results.append({
                        'NCM': row['NCM'],
                        'Descricao': row['Descricao'],
                        'Status': 'ERRO',
                        'NCM Sugerido': '',
                        'Justificativa': str(e)
                    })
            if i + self.batch_size < len(df_user):
                logging.info(f"Aguardando {self.wait_time} segundos antes do próximo lote...")
                time.sleep(self.wait_time)
        return pd.DataFrame(results)

    def generate_report(self, df_results: pd.DataFrame, df_ncm: pd.DataFrame, output_file: str):
        """Gera o relatório final em Excel com cor no status e descrição concatenada."""
        try:
            df_results['NCM'] = df_results['NCM'].astype(str).str.zfill(8)
            df_ncm['Codigo'] = df_ncm['Codigo'].astype(str).str.zfill(8)
            # Merge com a tabela oficial de NCMs
            df_final = df_results.merge(
                df_ncm,
                left_on='NCM',
                right_on='Codigo',
                how='left'
            )
            # Descrição concatenada
            df_final['Descrição Oficial'] = df_final['NCM'].apply(lambda x: self.buscar_descricao_concatenada(x, df_ncm))
            # Organiza as colunas
            df_final = df_final[[
                'NCM',
                'Descricao_x',  # do usuário
                'Descrição Oficial',
                'Status',
                'NCM Sugerido',
                'Justificativa'
            ]]
            df_final.columns = [
                'NCM Usuário',
                'Descrição Usuário',
                'Descrição Oficial',
                'Status',
                'NCM Sugerido',
                'Justificativa'
            ]
            # Salva o relatório
            df_final.to_excel(output_file, index=False)
            # Aplica cor na coluna Status
            self.aplicar_cor_status_excel(output_file)
            logging.info(f"Relatório gerado com sucesso: {output_file}")
        except Exception as e:
            logging.error(f"Erro ao gerar relatório: {str(e)}")
            raise

    def aplicar_cor_status_excel(self, output_file: str):
        """Aplica cor verde para CORRETO e vermelha para INCORRETO na coluna Status."""
        wb = load_workbook(output_file)
        ws = wb.active
        status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Status':
                status_col = idx
                break
        if status_col:
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
                for cell in row:
                    if str(cell.value).strip().upper() == 'CORRETO':
                        cell.fill = green
                    elif str(cell.value).strip().upper() == 'INCORRETO':
                        cell.fill = red
        wb.save(output_file)

def main():
    # Configurações
    API_KEY = ""
    INPUT_USER_FILE = "entrada/planilha_usuario.xlsx"
    INPUT_NCM_FILE = "entrada/Tabela_NCM_Vigente.xlsx"
    OUTPUT_FILE = "saida/relatorio_validacao_ncm.xlsx"
    os.makedirs("saida", exist_ok=True)
    validator = NCMValidator(API_KEY)
    try:
        df_user, df_ncm = validator.load_excel_files(INPUT_USER_FILE, INPUT_NCM_FILE)
        df_results = validator.process_in_batches(df_user)
        validator.generate_report(df_results, df_ncm, OUTPUT_FILE)
        logging.info("Processamento concluído com sucesso!")
    except Exception as e:
        logging.error(f"Erro durante o processamento: {str(e)}")
        raise

if __name__ == "__main__":
    main() 